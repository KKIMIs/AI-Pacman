import pygame
import sys
from settings import *
from pygame.math import Vector2 as vec
from class_player import *
from ghost import *
from red_ghost import *
from green_ghost import *
from blue_ghost import *
import genome_player
import genome_ghost
import numpy as np
from copy import *
from main_test import *
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side

USE_PREVIOUS_GENOME = True

class Generation():
    def __init__(self):
        # RULE!! : num_population % (num_best + num_children) = 0
        # RULE!! : num_children < 10
        self.genomes_p = []
        self.genomes_g = []
        self.best_genomes_p = []
        self.best_genomes_g = []
        self.num_population = 20
        self.num_best =5
        self.num_children =5
        self.prob_of_mutation = 0.4
        self.pre_genome_p = None
        self.pre_genome_g = None

    def set_initial_genomes(self):
        if USE_PREVIOUS_GENOME:

            self.genomes_p = [p_genome() for _ in range(self.num_population-1)]
            self.genomes_g = [GhostGenome() for _ in range(self.num_population-1)]
            p_previous_genome, g_previous_genome = p_genome() , GhostGenome()

            p_previous_genome.w1 = np.loadtxt('./record/p_genomew1_saved.cvs',delimiter=',')
            p_previous_genome.w2 = np.loadtxt('./record/p_genomew2_saved.cvs',delimiter=',')
            p_previous_genome.w3 = np.loadtxt('./record/p_genomew3_saved.cvs',delimiter=',')
            p_previous_genome.w4 = np.loadtxt('./record/p_genomew4_saved.cvs',delimiter=',')
            g_previous_genome.w1 = np.loadtxt('./record/g_genomew1_saved.cvs',delimiter=',')
            g_previous_genome.w2 = np.loadtxt('./record/g_genomew2_saved.cvs',delimiter=',')
            g_previous_genome.w3 = np.loadtxt('./record/g_genomew3_saved.cvs',delimiter=',')
            g_previous_genome.w4 = np.loadtxt('./record/g_genomew4_saved.cvs',delimiter=',')

            self.genomes_p.insert(0, p_previous_genome)
            self.genomes_g.insert(0, g_previous_genome)

        else:
            self.genomes_p = [p_genome() for _ in range(self.num_population)]
            self.genomes_g = [GhostGenome() for _ in range(self.num_population)]

    def keep_best_genomes(self):
        # 적합도가 높은 순으로 정렬
        if self.best_genomes_p is not None:
            self.genomes_p.extend(self.best_genomes_p)
        if self.best_genomes_g is not None:
            self.genomes_g.extend(self.best_genomes_g)

        fitness_p, fitness_g = [], []
        set_genomes_p, set_genomes_g = [], []

        for genome in self.genomes_p:
            fitness_p.append(genome.fitness)
        fitness_p = set(fitness_p)
        for genome in self.genomes_g:
            fitness_g.append(genome.fitness)
        fitness_g = set(fitness_g)

        for genome in self.genomes_p:
            if genome.fitness in fitness_p:
                set_genomes_p.append(genome)
                fitness_p.remove(genome.fitness)
        for genome in self.genomes_g:
            if genome.fitness in fitness_g:
                set_genomes_g.append(genome)
                fitness_g.remove(genome.fitness)

        if len(set_genomes_p) < self.num_best:
            num_shortage = self.num_best - len(set_genomes_p)
            for i in range(num_shortage):
                set_genomes_p.append(random.choice(set_genomes_p))
        if len(set_genomes_g) < self.num_best:
            num_shortage = self.num_best - len(set_genomes_g)
            for i in range(num_shortage):
                set_genomes_g.append(random.choice(set_genomes_g))

        set_genomes_p.sort(key=lambda x: x.fitness, reverse=True)

        #매 세대 임의로 유전자 삽입
        self.best_genomes_p = deepcopy(set_genomes_p[:self.num_best])
        # self.best_genomes_p.append(self.pre_genome_p)

        set_genomes_g.sort(key=lambda x: x.fitness, reverse=True)
        self.best_genomes_g = deepcopy(set_genomes_g[:self.num_best])
        # self.best_genomes_g = self.best_genomes_g + self.pre_genome_g

        # print('1. all genomes fitness = %s' % len(set_genomes_p))
        # print(','.join(str(genome.fitness) for genome in self.genomes_p))
        print('*****best genomes fitness*****')
        for i in range(self.num_best):
            print(i,":",self.best_genomes_p[i].fitness,"//",self.best_genomes_g[i].fitness)

    def cross_over(self, best_genomes, flag):
        for i in range(self.num_children):
            new_genome = deepcopy(best_genomes[0])
            a_genome = random.choice(best_genomes)
            b_genome = random.choice(best_genomes)

            # new_genome.w1.shape[1] == matrix 열 개수
            cut = random.randint(0, new_genome.w1.shape[1])
            # i 번째 자식의 첫번째~랜덤번째 유전자 = 상위 부모 개체의 유전자 일부
            new_genome.w1[i, :cut] = a_genome.w1[i, :cut]
            # i 번째 자식의 랜덤번째 유전자~마지막 유전자 = 상위 부모 개체의 유전자 일부
            new_genome.w1[i, cut:] = b_genome.w1[i, cut:]

            cut = random.randint(0, new_genome.w2.shape[1])
            new_genome.w2[i, :cut] = a_genome.w2[i, :cut]
            new_genome.w2[i, cut:] = b_genome.w2[i, cut:]

            cut = random.randint(0, new_genome.w3.shape[1])
            new_genome.w3[i, :cut] = a_genome.w3[i, :cut]
            new_genome.w3[i, cut:] = b_genome.w3[i, cut:]

            # 새로운 자식을 상위 개체 집단에 추가
            if flag == 'player':
                self.best_genomes_p.append(new_genome)
            elif flag == 'ghost':
                self.best_genomes_g.append(new_genome)


    def mutations(self, best_genomes, input_size, w1, w2, w3, output_size, flag):
        counter = 0
        for i in range(int(self.num_population / (self.num_best + self.num_children))):
            for bg in best_genomes:
                new_genome = deepcopy(bg)
                mean = 20  # 평균
                stddev = 10  # 표준편차

                # uniform(최소, 최대) 사이의 float 랜덤 반환
                if random.uniform(0, 1) < self.prob_of_mutation:
                    # np.random.normal(평균, 표준편차, 행렬or갯수) 정규분포 무작위 샘플 추출
                    new_genome.w1 += new_genome.w1 * np.random.normal(mean, stddev,
                                                                      size=(input_size, w1)) / 100 * np.random.randint(
                        -1, 2, (input_size, w1))
                if random.uniform(0, 1) < self.prob_of_mutation:
                    new_genome.w2 += new_genome.w2 * np.random.normal(mean, stddev,
                                                                      size=(w1, w2)) / 100 * np.random.randint(-1, 2,
                                                                                                               (w1, w2))

                if random.uniform(0, 1) < self.prob_of_mutation:
                    new_genome.w3 += new_genome.w3 * np.random.normal(mean, stddev,
                                                                      size=(w2, w3)) / 100 * np.random.randint(-1, 2,
                                                                                                               (w2, w3))

                if random.uniform(0, 1) < self.prob_of_mutation:
                    new_genome.w4 += new_genome.w4 * np.random.normal(mean, stddev,
                                                                      size=(w3, output_size)) / 100 * np.random.randint(
                        -1, 2,
                        (w3, output_size))

                if flag == 'player':
                    self.genomes_p.append(new_genome)
                elif flag == 'ghost':
                    self.genomes_g.append(new_genome)
                counter += 1


###################### Record(Exel) #############################
def set_up_sheet(sheet1, sheet2):
    sheet1['A1'].value = 'Generation'
    sheet1['B1'].value = 'Genome'
    sheet1['C1'].value = 'Player Fitness'
    sheet1['D1'].value = 'Ghost Fitness'
    sheet1['E1'].value = 'Play Time'

    sheet2['A1'].value = 'Generation'
    sheet2['B1'].value = 'Player Max Fitness'
    sheet2['C1'].value = 'Ghost Max Fitness'
    sheet2['D1'].value = 'Min Play Time'
    sheet2['E1'].value = 'Max Play Time'

    font = Font(size=12, bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    fill_lightgray = PatternFill(patternType='solid', fgColor=Color('D5D5D5'))
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

    for index in range(1, 6):
        sheet1.cell(row=1, column=index).font = font
        sheet2.cell(row=1, column=index).font = font
        sheet1.cell(row=1, column=index).alignment = align_center
        sheet2.cell(row=1, column=index).alignment = align_center
        sheet1.cell(row=1, column=index).fill = fill_lightgray
        sheet2.cell(row=1, column=index).fill = fill_lightgray
        sheet1.cell(row=1, column=index).border = border_thin
        sheet2.cell(row=1, column=index).border = border_thin

    sheet1.column_dimensions['A'].width = 15
    sheet1.column_dimensions['B'].width = 15
    sheet1.column_dimensions['C'].width = 20
    sheet1.column_dimensions['D'].width = 20
    sheet1.column_dimensions['E'].width = 15

    sheet2.column_dimensions['A'].width = 15
    sheet2.column_dimensions['B'].width = 20
    sheet2.column_dimensions['C'].width = 20
    sheet2.column_dimensions['D'].width = 15
    sheet2.column_dimensions['E'].width = 15

def record_best_genome(best_genomes,flag):
    if flag == "player":
        np.savetxt("./record/p_genomew1_saved.cvs",best_genomes.w1, delimiter=",")
        np.savetxt("./record/p_genomew2_saved.cvs",best_genomes.w2, delimiter=",")
        np.savetxt("./record/p_genomew3_saved.cvs",best_genomes.w3, delimiter=",")
        np.savetxt("./record/p_genomew4_saved.cvs",best_genomes.w4, delimiter=",")
    elif flag == "ghost":
        np.savetxt("./record/g_genomew1_saved.cvs",best_genomes.w1, delimiter=",")
        np.savetxt("./record/g_genomew2_saved.cvs",best_genomes.w2, delimiter=",")
        np.savetxt("./record/g_genomew3_saved.cvs",best_genomes.w3, delimiter=",")
        np.savetxt("./record/g_genomew4_saved.cvs",best_genomes.w4, delimiter=",")

####################### co-evolution start ######################
pygame.init()
icon = pygame.image.load('images/icon.png')
pygame.display.set_icon(icon)
pygame.display.set_caption('Pac-Man')
file = openpyxl.Workbook()
sheet1 = file.active
sheet1.title = 'Detailed Record'
sheet2 = file.create_sheet('Overall Record')
set_up_sheet(sheet1, sheet2)
align_center = Alignment(horizontal='center', vertical='center')

# generate 1st population
gen = Generation()
gen.set_initial_genomes()
num_gen = 0
row_index = 2
while True:
    num_gen += 1

    print("====================================================================")
    print("Generation: %s\n" % num_gen)

    gen_max_player_fitness = float('-inf')
    gen_max_ghost_fitness = float('-inf')
    max_play_time = float('-inf')
    min_play_time = float('inf')

    for i, gp in enumerate(gen.genomes_p):
        gg = gen.genomes_g[i]
        game = Game(Player, gp, gg, GreenGhost, RedGhost, PinkGhost, BlueGhost)
        player_fitness, ghost_fitness = game.run()

        play_time = round(game.play_time / 60, 2)
        gp.fitness, gg.fitness = player_fitness, ghost_fitness

        if gen_max_player_fitness < gp.fitness:
            gen_max_player_fitness = gp.fitness
        if gen_max_ghost_fitness < gg.fitness:
            gen_max_ghost_fitness = gg.fitness
        if max_play_time < play_time:
            max_play_time = play_time
        if min_play_time > play_time:
            min_play_time = play_time

        print('=====Genome #%4d-%2d, Player Fitness: %4.2f, Ghost Fitness: %4.2f, Play Time: %2.1f=====' % (num_gen, i + 1, gp.fitness, gg.fitness, play_time))
        sheet1.cell(row=row_index, column=1).value = num_gen
        sheet1.cell(row=row_index, column=1).alignment = align_center
        sheet1.cell(row=row_index, column=2).value = i + 1
        sheet1.cell(row=row_index, column=2).alignment = align_center
        sheet1.cell(row=row_index, column=3).value = gp.fitness
        sheet1.cell(row=row_index, column=3).alignment = align_center
        sheet1.cell(row=row_index, column=4).value = gg.fitness
        sheet1.cell(row=row_index, column=4).alignment = align_center
        sheet1.cell(row=row_index, column=5).value = play_time
        sheet1.cell(row=row_index, column=5).alignment = align_center
        row_index += 1

    gen.keep_best_genomes()
    record_best_genome(gen.best_genomes_p[0],"player")
    record_best_genome(gen.best_genomes_g[0],"ghost")
    gen.cross_over(gen.best_genomes_p, 'player')
    gen.cross_over(gen.best_genomes_g, 'ghost')
    gen.genomes_p, gen.genomes_g = [], []
    p_g, g_g  = p_genome(), GhostGenome()
    gen.mutations(gen.best_genomes_p, p_g.input_size, p_g.hidden_layer_size1, p_g.hidden_layer_size2, p_g.hidden_layer_size3, p_g.output_size, 'player')
    gen.mutations(gen.best_genomes_g, g_g.input_size, g_g.hidden_layer_size1, g_g.hidden_layer_size2, g_g.hidden_layer_size3, g_g.output_size, 'ghost')

    print("\nGeneration#%4d - Player Max Fitness: %4.2f, Ghost Max Fitness: %4.2f, Min Play Time: %2.1f, Max Play Time: %2.1f" % (num_gen, gen_max_player_fitness, gen_max_ghost_fitness, min_play_time, max_play_time))
    print("====================================================================\n\n")

    sheet2.cell(row=(num_gen + 1), column=1).value = num_gen
    sheet2.cell(row=(num_gen + 1), column=1).alignment = align_center

    sheet2.cell(row=(num_gen + 1), column=2).value = gen_max_player_fitness
    sheet2.cell(row=(num_gen + 1), column=2).alignment = align_center

    sheet2.cell(row=(num_gen + 1), column=3).value = gen_max_ghost_fitness
    sheet2.cell(row=(num_gen + 1), column=3).alignment = align_center

    sheet2.cell(row=(num_gen + 1), column=4).value = min_play_time
    sheet2.cell(row=(num_gen + 1), column=4).alignment = align_center

    sheet2.cell(row=(num_gen + 1), column=5).value = max_play_time
    sheet2.cell(row=(num_gen + 1), column=5).alignment = align_center

    file.save('./record/Record of co-evolution.xlsx')

        #genome_pre_using_settings
        # self.genomes_p = [p_genome() for _ in range(self.num_population-3)]
        # p_previous_genome = [p_genome() for _ in range(3)]
        #
        # for i in range(2):
            # if i%2 == 0:
            #     p_previous_genome[i].w1 = np.loadtxt('./record/p_g_sample0_w1_saved.cvs',delimiter=',')
            #     p_previous_genome[i].w2 = np.loadtxt('./record/p_g_sample0_w2_saved.cvs',delimiter=',')
            #     p_previous_genome[i].w3 = np.loadtxt('./record/p_g_sample0_w3_saved.cvs',delimiter=',')
            #     p_previous_genome[i].w4 = np.loadtxt('./record/p_g_sample0_w4_saved.cvs',delimiter=',')
            # else:
        #     target1 = "./record/p_g_sample" + str(i) + "_w1_saved.cvs"
        #     target2 = "./record/p_g_sample" + str(i) + "_w2_saved.cvs"
        #     target3 = "./record/p_g_sample" + str(i) + "_w3_saved.cvs"
        #     target4 = "./record/p_g_sample" + str(i) + "_w4_saved.cvs"
        #
        #     p_previous_genome[i].w1 = np.loadtxt(target1,delimiter=',')
        #     p_previous_genome[i].w2 = np.loadtxt(target2,delimiter=',')
        #     p_previous_genome[i].w3 = np.loadtxt(target3,delimiter=',')
        #     p_previous_genome[i].w4 = np.loadtxt(target4,delimiter=',')
        #
        # # self.pre_genome_p = p_previous_genome[0]
        # self.genomes_p =  p_previous_genome + self.genomes_p
        # # print(len(self.genomes_p))
        #
        # self.genomes_g = [GhostGenome() for _ in range(self.num_population-1)]
        # g_previous_genome = GhostGenome()
        #
        #
        # g_previous_genome.w1 = np.loadtxt('./record/g_genomew1_saved.cvs',delimiter=',')
        # g_previous_genome.w2 = np.loadtxt('./record/g_genomew2_saved.cvs',delimiter=',')
        # g_previous_genome.w3 = np.loadtxt('./record/g_genomew3_saved.cvs',delimiter=',')
        # g_previous_genome.w4 = np.loadtxt('./record/g_genomew4_saved.cvs',delimiter=',')
        #
        # self.genomes_g.insert(0, g_previous_genome)


        # for i in range(2):
        #     # if i%2 == 0:
        #     #     g_previous_genome[i].w1 = np.loadtxt('./record/g_g_sample0_w1_saved.cvs',delimiter=',')
        #     #     g_previous_genome[i].w2 = np.loadtxt('./record/g_g_sample0_w2_saved.cvs',delimiter=',')
        #     #     g_previous_genome[i].w3 = np.loadtxt('./record/g_g_sample0_w3_saved.cvs',delimiter=',')
        #     #     g_previous_genome[i].w4 = np.loadtxt('./record/g_g_sample0_w4_saved.cvs',delimiter=',')
        #     # else:
        #     target1 = "./record/g_g_sample" + str(i) + "_w1_saved.cvs"
        #     target2 = "./record/g_g_sample" + str(i) + "_w2_saved.cvs"
        #     target3 = "./record/g_g_sample" + str(i) + "_w3_saved.cvs"
        #     target4 = "./record/g_g_sample" + str(i) + "_w4_saved.cvs"
        #
        #     g_previous_genome[i].w1 = np.loadtxt(target1,delimiter=',')
        #     g_previous_genome[i].w2 = np.loadtxt(target2,delimiter=',')
        #     g_previous_genome[i].w3 = np.loadtxt(target3,delimiter=',')
        #     g_previous_genome[i].w4 = np.loadtxt(target4,delimiter=',')


        # self.pre_genome_g = g_previous_genome[:2]
        # self.genomes_g =  g_previous_genome + self.genomes_g
