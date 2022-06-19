import pygame
import sys
from settings import *
from pygame.math import Vector2 as vec
from main_test import *
from class_player import *
from ghost import *
from red_ghost import *
from green_ghost import *
from blue_ghost import *
from pink_ghost import *
from genome_player import *
import numpy as np
from copy import deepcopy
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side

USE_PREVIOUS_GENOME = True

class Generation():
	def __init__(self):
		#RULE!! : num_population % (num_best + num_children) = 0 이여야 한다.
		#RULE!! : num_children < 10
		self.genomes = []
		self.best_genomes = []
		self.num_population = 20
		self.num_best = 5
		self.num_children = 5
		self.prob_of_mutation = 0.5

	def set_initial_genomes(self):
		if USE_PREVIOUS_GENOME:
			self.genomes = [p_genome() for _ in range(self.num_population-1)]
			previous_genome = p_genome()
			previous_genome.w1 = np.loadtxt('./record/genomew1_saved.cvs',delimiter=',')
			previous_genome.w2 = np.loadtxt('./record/genomew2_saved.cvs',delimiter=',')
			previous_genome.w3 = np.loadtxt('./record/genomew3_saved.cvs',delimiter=',')
			previous_genome.w4 = np.loadtxt('./record/genomew4_saved.cvs',delimiter=',')
			self.genomes.insert(0, previous_genome)

		else:
			self.genomes = [p_genome() for _ in range(self.num_population)]

	def keep_best_genomes(self):
		if self.best_genomes is not None:
			self.genomes.extend(self.best_genomes)
			#메인 게놈에 추가

		# 적합도가 높은 순으로 정렬
		self.genomes.sort(key=lambda x: x.fitness, reverse=True)

		# 상위 num_best개 선택
		self.best_genomes = deepcopy(self.genomes[:self.num_best])

	def cross_over(self, best_genomes):
		for i in range(self.num_children):
			new_genome = deepcopy(best_genomes[0])
			a_genome = random.choice(best_genomes)
			b_genome = random.choice(best_genomes)

			# new_genome.w1.shape[1] == matrix 열 개수
			cut = random.randint(0, new_genome.w1.shape[1])
			# i 번째 자식의 첫번째~랜덤번째 유전자 = 상위 부모 개체의 유전자 일부
			new_genome.w1[i, :cut] = a_genome.w1[i, :cut]
			# i 번째 자식의 랜덤번째 유전자~마지막 유전자 = 상위 부모 개체의 유전자 일부
			new_genome.w1[i, cut:] = b_genome.w1[i, cut:]

			cut = random.randint(0, new_genome.w2.shape[1])
			new_genome.w2[i, :cut] = a_genome.w2[i, :cut]
			new_genome.w2[i, cut:] = b_genome.w2[i, cut:]

			cut = random.randint(0, new_genome.w3.shape[1])
			new_genome.w3[i, :cut] = a_genome.w3[i, :cut]
			new_genome.w3[i, cut:] = b_genome.w3[i, cut:]

			cut = random.randint(0, new_genome.w4.shape[1])
			new_genome.w4[i, :cut] = a_genome.w4[i, :cut]
			new_genome.w4[i, cut:] = b_genome.w4[i, cut:]

			# 새로운 자식을 상위 개체 집단에 추가
			self.best_genomes.append(new_genome)

	def mutations(self, best_genomes):
		counter = 0
		for i in range(int(self.num_population / (self.num_best + self.num_children))):
			for bg in best_genomes:
				new_genome = deepcopy(bg)
				mean = 20   #평균
				stddev = 10   #표준편차

				# uniform(최소, 최대) 사이의 float 랜덤 반환
				if random.uniform(0, 1) < self.prob_of_mutation:
					# np.random.normal(평균, 표준편차, 행렬or갯수) 정규분포 무작위 샘플 추출
					new_genome.w1 += new_genome.w1 * np.random.normal(mean, stddev, size=(9, 15)) / 100 * np.random.randint(-1, 2, (9, 15))

				if random.uniform(0, 1) < self.prob_of_mutation:
					new_genome.w2 += new_genome.w2 * np.random.normal(mean, stddev, size=(15, 20)) / 100 * np.random.randint(-1, 2, (15, 20))

				if random.uniform(0, 1) < self.prob_of_mutation:
					new_genome.w3 += new_genome.w3 * np.random.normal(mean, stddev, size=(20, 10)) / 100 * np.random.randint(-1, 2, (20, 10))

				if random.uniform(0, 1) < self.prob_of_mutation:
					new_genome.w4 += new_genome.w4 * np.random.normal(mean, stddev, size=(10, 4)) / 100 * np.random.randint(-1, 2, (10, 4))

				self.genomes.append(new_genome)
				counter +=1

def set_up_sheet(sheet1, sheet2):
	sheet1['A1'].value = 'Generation'
	sheet1['B1'].value = 'Genome'
	sheet1['C1'].value = 'Player Fitness'
	sheet1['D1'].value = 'Play Time'

	sheet2['A1'].value = 'Generation'
	sheet2['B1'].value = 'Player Max fitness'
	sheet2['C1'].value = 'Min Play Time'
	sheet2['D1'].value = 'Max Play Time'

	font = Font(size=12, bold=True)
	align_center = Alignment(horizontal='center', vertical='center')
	fill_lightgray = PatternFill(patternType='solid', fgColor=Color('D5D5D5'))
	border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

	for index in range(1, 5):
		sheet1.cell(row=1, column=index).font = font
		sheet1.cell(row=1, column=index).alignment = align_center
		sheet1.cell(row=1, column=index).fill = fill_lightgray
		sheet1.cell(row=1, column=index).border = border_thin

	for index in range(1, 5):
		sheet2.cell(row=1, column=index).font = font
		sheet2.cell(row=1, column=index).alignment = align_center
		sheet2.cell(row=1, column=index).fill = fill_lightgray
		sheet2.cell(row=1, column=index).border = border_thin

	sheet1.column_dimensions['A'].width = 15
	sheet1.column_dimensions['B'].width = 15
	sheet1.column_dimensions['C'].width = 20
	sheet1.column_dimensions['D'].width = 15

	sheet2.column_dimensions['A'].width = 15
	sheet2.column_dimensions['B'].width = 20
	sheet2.column_dimensions['C'].width = 15
	sheet2.column_dimensions['D'].width = 15

def record_best_genome(best_genomes):
	np.savetxt("./record/genomew1_saved.cvs",best_genomes.w1, delimiter=",")
	np.savetxt("./record/genomew2_saved.cvs",best_genomes.w2, delimiter=",")
	np.savetxt("./record/genomew3_saved.cvs",best_genomes.w3, delimiter=",")
	np.savetxt("./record/genomew4_saved.cvs",best_genomes.w4, delimiter=",")


pygame.init()
icon = pygame.image.load('images/icon.png')
pygame.display.set_icon(icon)
#s = pygame.display.set_mode((SCREEN_WIDTH*2, SCREEN_HEIGHT))-->Game.screen
pygame.display.set_caption('Pac-Man')

file = openpyxl.Workbook()
sheet1 = file.active
sheet1.title = 'player'
sheet2 = file.create_sheet('max fitness')
set_up_sheet(sheet1, sheet2)
align_center = Alignment(horizontal='center', vertical='center')

# generate 1st population
gen = Generation()
gen.set_initial_genomes()

num_gen = 0
row_index = 2

while True:
	num_gen += 1

	print("====================================================================")
	print("Generation: %s\n" % num_gen)

	gen_max_player_fitness = float('-inf')
	max_play_time = float('-inf')
	min_play_time = float('inf')

	for i, genome in enumerate(gen.genomes):

		game = Game(Player, genome, GreenGhost, RedGhost, PinkGhost, BlueGhost)
		player_fitness = game.run()

		play_time = round(game.play_time/60,2)
		genome.fitness = player_fitness

		print('=====Genome #%4d-%2d, Player Fitness: %4.2f, Play Time: %2.1f=====' % (num_gen, i+1, genome.fitness,play_time))
		sheet1.cell(row=row_index, column=1).value = num_gen
		sheet1.cell(row=row_index, column=2).value = i+1
		sheet1.cell(row=row_index, column=3).value = genome.fitness
		sheet1.cell(row=row_index, column=4).value = play_time
		sheet1.cell(row=row_index, column=1).alignment = align_center
		sheet1.cell(row=row_index, column=2).alignment = align_center
		sheet1.cell(row=row_index, column=3).alignment = align_center
		sheet1.cell(row=row_index, column=4).alignment = align_center
		row_index += 1

		if gen_max_player_fitness < genome.fitness:
			gen_max_player_fitness = genome.fitness
		if max_play_time < play_time:
			max_play_time = play_time
		if min_play_time > play_time:
			min_play_time = play_time

	gen.keep_best_genomes()
	record_best_genome(gen.best_genomes[0])
	gen.cross_over(gen.best_genomes)
	gen.genomes = []
	gen.mutations(gen.best_genomes)

	print("Generation#%s - Player Max Fitness: %s" % (num_gen, gen_max_player_fitness))
	sheet2.cell(row=(num_gen+1), column=1).value = num_gen
	sheet2.cell(row=(num_gen+1), column=2).value = gen_max_player_fitness
	sheet2.cell(row=(num_gen + 1), column=3).value = min_play_time
	sheet2.cell(row=(num_gen + 1), column=4).value = max_play_time

	sheet2.cell(row=(num_gen+1), column=1).alignment = align_center
	sheet2.cell(row=(num_gen+1), column=2).alignment = align_center
	sheet2.cell(row=(num_gen + 1), column=3).alignment = align_center
	sheet2.cell(row=(num_gen + 1), column=4).alignment = align_center

	file.save('./record/Record of evolution.xlsx')
record.close()
